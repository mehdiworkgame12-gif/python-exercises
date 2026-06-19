a=[1,2,3]
print(a)
type(a)
print(type(a))
s1= {1,2,3,4}
s2= {1,2}
print(s1.union(s2))
s1.intersection(s2)
print(s1|s2)
def function2():
    s = 'kdyhfg@#@!'
    target='@#'
    return(set(s).intersection(set(target)))
function2()
print(function2())
def p2p3(x):
    p2 = x**2
    p3= x**3
    return[p2,p3]
print(p2p3(10))
print(p2p3(11))
def p1p2(f):
    return f['x']**f['y']
print(p1p2({'x':2,'y':3}))
def p1p2(x):
    return x[0]**x[1]
print(p1p2([2,3]))
def p1p2(x,y):
    return x**y
print(p1p2(10,3))
def function2(s,target):
    return(set(s).intersection(set(target)))
a=function2('hdfbv!@','@!')
print(a)
def zarbkon(x,y):
    return x*y
print(zarbkon('2',3))
print(zarbkon(2,3))
def jamzarb(a,b) ->dict:
    return {'jam':a+b,'zarb':a*b}
print(jamzarb(2,3))
def jamzarb(a,b) ->dict:
    return {'jam':a+b,'zarb':a+b}
print(jamzarb('2','3'))
#جمع رشته داریم ضرب نداریم منظور ضرب 2 رشته نداریم
#تو مسعله قبل اگر به جای جمع ضرب بود کار نمیکرد
#'hi'*3=hihihi
a='mehdi'
b='king'
print(a+b)
#پایان return در تابع

#raise در پایتون ینی عمدا یک ارور خاص مشخص کنی
#پایان چک کردن تایپ در تابع
def mehdi(a:float,b:float=10.0):
    return a*b
print(mehdi(2,3))
#در تمرین قبل اگر به جای خط اخر بجای b هیچی نمیذاشتیم چون مقدار دیفالت b 10 بود بازم 2 ضرب 10 را نشان میداد
#پایان وردی های defaultدر تابع
for i in range(3):
    print('salam',end=' ')
#end ینی نرو خط بعدی سلام را چاپ کن مثلا اینجا گفتم فاصله بذار
#پایان مثال با تابع
def zojfard(a:list)->dict:
    if not isinstance(a,list):
        raise TypeError('bayad list bashad')
    result={'zoj':[],'fard':[]}
    for i in a:
        if i%2==0:
            result['zoj'].append(i)
        else:
            result['fard'].append(i)
    return result
print(zojfard([1,2,3,4,5]))
#تمرین پایین برای نمایش زوج یا قرد بودن عدده
def adadaval(n:int)-> bool:
    for i in range(2,n):
        if n%i==0:
            return False
    return True
print(adadaval(13))
#در اینجا میتواینم در خط دوم بجای n بنویسیم:int(n**0.5)+1
#پایان تابع پیدا کردن عدد اول در یک لیست

def zojfard(a:int)->dict:
    result={'zoj':[],'fard':[],'aval':[]}
    for i in a:
        if i%2==0:
            result['zoj'].append(i)
        else:
            result['fard'].append(i)
        if adadaval(i):
            result['aval'].append(i)
    return result
print(zojfard([1,2,3,4,5,7]))
#در اینجا از فرام مهدی کینگ اون تابع مسعله قبلیو استفاده کردیم
#شروع امکانات پای چارم برای دی باگینگ
a=1
b=23
c='ab'
print(str(b)+c)
#تمرین اول این فصل پایین
def maxminl(a):
    result={'max':max(a),'min':min(a),'len':len(a)}
    return result
a=[1,2,3,4,7,43]
print(maxminl(a))
#تمرین دوم این فصل
def miangin(data):
    result={}
    for key,value in data.items():
        avg=sum(value)/len(value)
        result[key]=avg
    return result

data={'a1':[1,2,3],'a2':[4,5,6],'a3':[7,8,9]}

print(miangin(data))
#تمرین سوم
def eshtrakoejtema(data):
    sets = list(data.values())
    intersection=sets[0]
    for i in sets[1:]:
        intersection= intersection & i
    union=sets[0]
    for i in sets[1:]:
        union=union | i
    return {'intersection':intersection,'union':union}
data={'s1':{1,2,3},'s2':{3,4,5,6}}
print(eshtrakoejtema(data))

#تمرین چهارم
def max2(a,b):
    if a>b:
        return a
    else:
        return b
def max3(a,b,c):
    return max2(max2(a,b),c)
print(max3(10,5,7))

#شروع فصل جدید iterator
a=(1,2,3,'b')
a=iter(a)
print(next(a))
print(next(a))
print(next(a))
a_iter=iter(a)
print(type(a_iter))

#آغاز generator
def my_generator():
    X = 100

    yield x

    x *= 10

    yield x

def zoj_generator(a:list):
    for i in a:
        if i%2==0:
            yield i
a=[1,2,3,4]
for i in zoj_generator(a):
    print(i)

def zoj_non_generator(a:list):
    zoj=[]
    for i in a:
        if i%2==0:
            zoj.append(i)
    return zoj

a=[1,2,3,4,5,6]
[i for i in a if i%2==0 if i>2]
a=[1,2,3,4,5,6]
[i if i%2==0 else -i for i in a]
#اغاز  توابع توو در توو
print([i for i in a if i%2==0 if i>2])
a=[1,2,3,4,5,6]
# اغاز توابع تووو در تووو
def f1():
    return f2(2)
def f2(x):
    return 3*x*f3(5)
def f3(a):
    return 2*a
print(f1())

def f1():
    return f2(2)
def f2(x):
    return 3*x*f3(5)
def f3(a):
    return 2*a
def outer_f():
    def inner_f():
        print('salam')
print(outer_f())

def add(n):
    def inner_add():
        return n+1
    return inner_add()
print(add(5))

def tolidtavan(tavan):
    def tolidmabna(mabna):
        return tavan**mabna
    return tolidmabna
tavan4=tolidtavan(4)
print(tavan4(2))


def rec_func(a):
    if a>10:
        print(a)

    else:
        return
#اغاز توابع فاکتوریل
#اغاز تابع فاکتوریل
#n!=(n-1)! for example 5!=5*4!
def factorial(n):
    if n==0:
        return 1
    return n*factorial(n-1)
print(factorial(5))
#اغاز ورودی *args kwargs*
def avarage(a,b,c):
    return (a+b+c)/3
print(avarage(3,4,5))

def avrage(*args):
    for x in args:
        print(x)

def avrage(*args):
    for x in args:
        if not isinstance(x,int):
            raise TypeError
    return sum(args)/len(args)

def avrage(*args):
    sum=0
    counter=0
    for x in args:
        if not isinstance(x,int,float):
            return 'Erorr'
        sum+=x
        counter+=1
    return sum/counter

def print_kw(**kwargs):
    print(kwargs)
print_kw(x=10,u=56)
#به شکل دیکشنری میاره و حتما نقدار و اینا باید بدی
def  avrage(**kwargs):
    return sum(kwargs.values())/len(kwargs.values())
print(avrage(x=10,u=12,b=56))

def  avrage(**kwargs):
    sum=0
    counter=0
    for k ,v in kwargs.items():
        sum+=v
        counter+=1
    return sum/counter
print(avrage(x=10,u=12,b=58))
#جایگذاری متغییر در رشته
a='mehdi'
print('name',a)
print('name:',a)
a='aqamehdi'
print('esme in karbar',a,'hast')
s='esme in karbar %s hast'%a
print(s)
s='sen %s %d hast' %(a,24)
print(s)
#توی پایتون کنسول چیزای مهمی نوشتم
#اغاز فصصصلل 5 کلاس و برنامه نویسی شی گرا
a='mehdiking','reza'.split(',')
print(a)
a=[1,2,3,4]
a.reverse()
print(a)

class Rectangle:
    pass
#r1=Rectangle(length=20,width=10)
#r2=Rectangle(length=200,width=10)
#r1.area()->200
#r2.perimeter->2*(20+10)
#اغاز مفهوم متدد

class Rectangle:
    def area(self):
        return width*height
#r.area()............Rectangle.area(r)
class Rectangle:
    def area(self):
        return self.width*self.height

#magic methods
class Rectangle:
    def __init__(self,width,height):
        self.width=width
        self.height=height
    def area(self):
        return self.width*self.height
    def __repr__(self) -> str:
        return f'Rectangle({self.length},{self.width})'
    def area(self):
        self.area_cal_counter+=1
        return self.length*self.width
    def area_factor(self,factor):
        return self.width*self.length*factor
    def __eq__(self,other):
        return self.area()==other.area()
    def __gt__(self,other):
        return self.area()>other.area()
    def  perimeter(self):
        return 2*(self.length+self.width)
    def area_factor(self,x):
        return (self.aera()*factor(x))
    def factor(a):
        return a**2
    def reducer(x):
        return x**0.1
    def reducer(Rectangle):
        return Rectangle.perimeter()*0.1

class Point:
    def __init__(self,x,y):
        self.x=x
        self.y=y
        self.size=0
    def __repr__(self):
        return f' point with x:{self.x},y:{self.y}'
    def  setsize(self,s):
        self.size=s

class color:
    def __init__(self,r,g,b):
        self.red=r
        self.green=g
        self.blue=b
    def __repr__(self):
        return f' color with r:{self.red},g:{self.green},b:{self.blue}'


class car:
    def __init__(self,model,year,color):
        self.model=model
        self.year=year
        self.color=color
        self.heater=Heater()
    def break_(self):
        print(f'khodro{self.model} tormoz kard')

#تمرین دوم
class Rectangle:
    def __init__(self,width,height):
        self.width=width
        self.height=height
    def area(self):
        return self.width*self.height
    def perimeter(self):
        return 2*(self.width+self.height)
    def __str__(self):
        return (f'Rectangle({self.width},{self.height})')
    def __eq__(self,other):
        return self.area()==other.area()
    def __gt__(self,other):
        return self.perimeter()>other.perimeter()
class square(Rectangle):
    def __init__(self,width):
        super().__init__(width,width)
    def __str__(self):
        return (f'square ({self.width},{self.height})')
r1=Rectangle(10,20)
print(r1)
print(r1.area())
print(r1.perimeter())
r2=Rectangle(10,20)
print(r1==r2)
print(r1>r2)
s=square(5)
print(s)
#اغاز فصل 6 مطالب پبشرفته در برنامه نویسی شی گرا
#ارث بری
class Employee:
    def __init__(self,lname,age,fname,idcart):
        self.age=age
        self.idcartr=idcart
        self.lname=lname
        self.fname=fname
    def access(self):
        print('general access')
class Itadmin(Employee):
    def __init__(self,lname,age,fname,idcart):
        super().__init__(fname,lname,idcart)
    def access(self):
        print('It')
    def print_name(self):
        print(self.name)
#تمرین بالایی خیلی مهمه عکس گرفتم با ایفون تاریخ 6 june ساعت 18:34
# issubclass برای اینکه ببینیم یک کلاس فرزند یا والد ذیگری هست یا ن
#شروه مفهوم انتزاغ abstact
from abc import ABC ,abstractmethod
class Animal(ABC):
    @property
    @abstractmethod
    def age(self):
        pass
    @abstractmethod
    def run(self):
        pass
    def walk(self):
        pass
class Human(Animal):
    @property
    def age(self):
        return 20
    def run(self):
        print('the Human is running')
#polymorphismشروع موضوع چند ریختی چند ریختی
class Circle:
    def __init__(self,radius):
        self.radius=radius

    def area(self):
        return self.radius**2*3.14
class square:
    def __init__(self,s):
        self.s=s
    def area(self):
        return self.s**2
#اغاز encapsulation کپسوله سازی
class ITManager:
    def __init__(self):
        self.__IT_manager_files = 'ert'
        self._IT_manager_file='abc'
    def print_files(self):
        print('public files')
    def __print_files(self):
        print('private files')
    def access_for_ceo(self):
        print(self.__IT_manager_file)
class ITadmin(ITManager):
    def __init__(self):
        ITManager.__init__(self)
        self.id=1234
#شروع class method and static method
class printer:
    def print_salam(self):
        print('salam')
    @classmethod
    def class_print_salam(cls):
        print('salam class')
    @staticmethod
    def static_print_salam():
        print('salam static')
#شروع فصل 7 اخر..........
#خواندن و نکات فایل csv
with open('data01.txt')as f:
    lines=f.readlines()

#ماژول سی اس وی
import csv
with open('data03.csv',mode='r')as f:
    content= csv.reader(f,delimiter='\t')
    for row in content:
        print(row)
print(content)
#متد رایت csv
header=['ostan','pop','paytakht']
data=[['fars',4.8,'shiraz',],['bushehr',1.5,'bushehr']]
with open('data04.csv',mode='w',newline='')as f:
    writer=csv.writer(f)
    writer.writerow(header)
    writer.writerows(data)
#jsonفایل
import json
d={'a':100,'b':200}
with open('sample00.json','w') as f:
    json.dump(d,f)
with open('sample00.json','r') as f:
    d2=json.load(f)
#method openpyxl
import openpyxl
import pandas as pd
wb=openpyxl.load_workbook('Book1.xlsx')
sheet1=wb['Sheet1']
col2=[]
for i in range(1,sheet1.max_row):
    col2.append(sheet1.cell(row=i,column=1).value)
wb2= openpyxl.Workbook()
s0=wb2.active
s0['a1']='name'
s0['b1']='year'
names=['apl','MSFT','google']
s0.cell(row=2,column=2).value=2002
for i,name in enumerate(names):
    s0.cell(row=i+ 2, column=1).value=name
wb2.save('write_book.xlsx')
s2=wb2.create_sheet('s2')
s2['a1']=1000
wb2.save('write_book.xlsx')
df=pd.read_excel('Book1.xlsx',sheet_name=0)
#pickle فایل
#pickle یک فایل باینری است
import pickle
d={'a':100}
with open('data.pickle','wb')as f:
    pickle.dump(d,f)
with open('data.pickle','rb')as f:
    d2=pickle.load(f)
import pickle
class car:
    def __init__(self,color,year):
        self.color=color
        self.year=year
        self.aaa=[1,2,4,5]

c1=car('b',2012)
with open('c1.pickle','wb')as f:
    pickle.dump(c1,f)
with open('c1.pickle','rb')as f:
    c1_read=pickle.load(f)
#try-except
print('welcome to my app')
x=float(input('x:'))
y=float(input('y:'))
print('x/y:',x/y)

print('welcome to my app')
while True:
    x=input('x:')
    if x=='q':
        break
    else:
        x=float(x)
    y=float(input('y:'))
    print('x/y:',x/y)
#ecept وثتی مینویسی ینی به جز
#zerodivisionEror
print('welcome to my app')
while True:
    x = input('x:')
    if x == 'q':
        break
    else:
        x = float(x)
    y = float(input('y:'))
    try:
        x / y
    except ZeroDivisionError:
        print('y can not be zero')
    except:
        print('something is wrong')
    else:
        print('x/y:', x / y)
#کار با فایل ها
import pandas as pd
files=['data04.csv']
res=[]
for file in files:
    try:
       df = pd.read_csv(file)
       print (df.iloc[0,0])
       res.append(df.iloc[0,0])
    except FileNotFoundError:
        print(f'{file}is not here')

#امکانات ماژول os برای کاربا فایل ها
import os
os.mkdir()
os.chdir()
os.getcwd()+'\\csvfiles'
newpath=os.getcwd()+'\\csvfiles'
import pandas as pd
pd.read_csv(newpath+'\\data04.csv')
os.getcwd()+os.sep+'csvfiles'
#os.sep برای // استقاده میشود
os.listdir()
#متد بالا همه رو میاره لیست ها و...
os.listdir(os.path.join(os.getcwd(),'csvfiles'))
#نکته بالا ینی هرچی در قایل csv هست برام بیار
for file in os.listdir():
    print(file)
os.listdir()[1]
os.listdir()[2]
os.listdir()[3]
os.listdir()[1].split(',')
for file in os.listdir():
    if file.split('.')[-1]=='csv':
        print(file)


os.path.isfile('data04.csv')
#true
os.path.isdir(os.listdir()[4])

for i,j,k in os.walk(os.getcwd()):
    print(i,j,k)
#پایان پایتون مقدماتی تبرررررررررررییککک
x=10
print(x*2)






